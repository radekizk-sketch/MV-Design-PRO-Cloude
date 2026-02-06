"""
XLSX Network Importer — Import sieci SN z arkuszy Excel

Format arkusza (uzgodniony z operatorami sieci):
- Arkusz "Szyny": id, nazwa, napięcie_kV
- Arkusz "Linie": id, szyna_pocz, szyna_kon, typ, dlugość_km, R_ohm_km, X_ohm_km, B_uS_km
- Arkusz "Trafo": id, szyna_HV, szyna_LV, Sn_MVA, uk_pct, Pk_kW, grupa
- Arkusz "Źródła": id, szyna, typ, Sk_MVA, RX_ratio
- Arkusz "Odbiory": id, szyna, P_MW, Q_Mvar

PRINCIPLES:
- Validates all data before creating network
- Polish error messages for user display
- Returns structured validation report
- Creates NetworkGraph compatible with existing solvers
"""
from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from typing import Any

from network_model.core.graph import NetworkGraph
from network_model.core.node import Node, NodeType
from network_model.core.branch import LineBranch, TransformerBranch, BranchType


class XlsxValidationError(Exception):
    """Validation error during XLSX import."""

    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__(f"Błędy walidacji importu XLSX: {len(errors)} błędów")


@dataclass
class XlsxImportResult:
    """Result of XLSX import operation."""

    success: bool
    graph: NetworkGraph | None
    bus_count: int = 0
    branch_count: int = 0
    source_count: int = 0
    load_count: int = 0
    trafo_count: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "bus_count": self.bus_count,
            "branch_count": self.branch_count,
            "source_count": self.source_count,
            "load_count": self.load_count,
            "trafo_count": self.trafo_count,
            "warnings": self.warnings,
            "errors": self.errors,
        }


class XlsxNetworkImporter:
    """
    Importer sieci SN z arkuszy Excel.

    Usage:
        importer = XlsxNetworkImporter()
        result = importer.import_from_bytes(xlsx_bytes)
        if result.success:
            graph = result.graph
    """

    REQUIRED_SHEETS = {"Szyny", "Linie"}
    OPTIONAL_SHEETS = {"Trafo", "Źródła", "Odbiory"}

    # Column mappings for each sheet
    BUS_COLUMNS = {"id": str, "nazwa": str, "napięcie_kV": float}
    LINE_COLUMNS = {
        "id": str,
        "szyna_pocz": str,
        "szyna_kon": str,
        "typ": str,
        "długość_km": float,
        "R_ohm_km": float,
        "X_ohm_km": float,
    }
    TRAFO_COLUMNS = {
        "id": str,
        "szyna_HV": str,
        "szyna_LV": str,
        "Sn_MVA": float,
        "uk_pct": float,
    }
    SOURCE_COLUMNS = {
        "id": str,
        "szyna": str,
        "typ": str,
        "Sk_MVA": float,
        "RX_ratio": float,
    }
    LOAD_COLUMNS = {
        "id": str,
        "szyna": str,
        "P_MW": float,
        "Q_Mvar": float,
    }

    def import_from_bytes(self, data: bytes) -> XlsxImportResult:
        """Import network from XLSX file bytes."""
        try:
            import openpyxl
        except ImportError:
            return XlsxImportResult(
                success=False,
                graph=None,
                errors=["Brak biblioteki openpyxl — zainstaluj: pip install openpyxl"],
            )

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True
            )
        except Exception as e:
            return XlsxImportResult(
                success=False,
                graph=None,
                errors=[f"Nie można otworzyć pliku XLSX: {e}"],
            )

        errors: list[str] = []
        warnings: list[str] = []

        # Check required sheets
        sheet_names = set(wb.sheetnames)
        for req in self.REQUIRED_SHEETS:
            if req not in sheet_names:
                errors.append(f"Brak wymaganego arkusza: '{req}'")

        if errors:
            return XlsxImportResult(success=False, graph=None, errors=errors)

        # Parse sheets
        buses = self._parse_sheet(wb["Szyny"], self.BUS_COLUMNS, "Szyny", errors)
        lines = self._parse_sheet(wb["Linie"], self.LINE_COLUMNS, "Linie", errors)

        trafos: list[dict[str, Any]] = []
        if "Trafo" in sheet_names:
            trafos = self._parse_sheet(
                wb["Trafo"], self.TRAFO_COLUMNS, "Trafo", errors
            )

        sources: list[dict[str, Any]] = []
        if "Źródła" in sheet_names:
            sources = self._parse_sheet(
                wb["Źródła"], self.SOURCE_COLUMNS, "Źródła", errors
            )

        loads: list[dict[str, Any]] = []
        if "Odbiory" in sheet_names:
            loads = self._parse_sheet(
                wb["Odbiory"], self.LOAD_COLUMNS, "Odbiory", errors
            )

        if errors:
            return XlsxImportResult(success=False, graph=None, errors=errors)

        # Validate cross-references
        bus_ids = {b["id"] for b in buses}

        for i, line in enumerate(lines):
            if line["szyna_pocz"] not in bus_ids:
                errors.append(
                    f"Linie wiersz {i + 2}: szyna_pocz '{line['szyna_pocz']}' "
                    f"nie istnieje w arkuszu Szyny"
                )
            if line["szyna_kon"] not in bus_ids:
                errors.append(
                    f"Linie wiersz {i + 2}: szyna_kon '{line['szyna_kon']}' "
                    f"nie istnieje w arkuszu Szyny"
                )

        for i, trafo in enumerate(trafos):
            if trafo["szyna_HV"] not in bus_ids:
                errors.append(
                    f"Trafo wiersz {i + 2}: szyna_HV '{trafo['szyna_HV']}' "
                    f"nie istnieje w arkuszu Szyny"
                )
            if trafo["szyna_LV"] not in bus_ids:
                errors.append(
                    f"Trafo wiersz {i + 2}: szyna_LV '{trafo['szyna_LV']}' "
                    f"nie istnieje w arkuszu Szyny"
                )

        for i, src in enumerate(sources):
            if src["szyna"] not in bus_ids:
                errors.append(
                    f"Źródła wiersz {i + 2}: szyna '{src['szyna']}' "
                    f"nie istnieje w arkuszu Szyny"
                )

        for i, load in enumerate(loads):
            if load["szyna"] not in bus_ids:
                errors.append(
                    f"Odbiory wiersz {i + 2}: szyna '{load['szyna']}' "
                    f"nie istnieje w arkuszu Szyny"
                )

        # Check for duplicate IDs
        all_bus_ids = [b["id"] for b in buses]
        if len(all_bus_ids) != len(set(all_bus_ids)):
            errors.append("Duplikaty ID w arkuszu Szyny")

        all_line_ids = [line["id"] for line in lines]
        if len(all_line_ids) != len(set(all_line_ids)):
            errors.append("Duplikaty ID w arkuszu Linie")

        # Validate numeric values
        for i, bus in enumerate(buses):
            if bus["napięcie_kV"] <= 0:
                errors.append(f"Szyny wiersz {i + 2}: napięcie_kV musi być > 0")

        for i, line in enumerate(lines):
            if line["długość_km"] <= 0:
                errors.append(f"Linie wiersz {i + 2}: długość_km musi być > 0")
            if line["R_ohm_km"] < 0:
                errors.append(f"Linie wiersz {i + 2}: R_ohm_km musi być >= 0")
            if line["X_ohm_km"] < 0:
                errors.append(f"Linie wiersz {i + 2}: X_ohm_km musi być >= 0")

        if errors:
            return XlsxImportResult(success=False, graph=None, errors=errors)

        # Build NetworkGraph
        graph = self._build_graph(buses, lines, trafos, sources, loads, warnings)

        return XlsxImportResult(
            success=True,
            graph=graph,
            bus_count=len(buses),
            branch_count=len(lines) + len(trafos),
            source_count=len(sources),
            load_count=len(loads),
            trafo_count=len(trafos),
            warnings=warnings,
        )

    def import_from_dict(
        self, data: dict[str, list[dict[str, Any]]]
    ) -> XlsxImportResult:
        """
        Import network from dict of sheet data (for testing without openpyxl).

        data = {
            "Szyny": [{"id": "B1", "nazwa": "Szyna główna", "napięcie_kV": 15.0}, ...],
            "Linie": [...],
            ...
        }
        """
        errors: list[str] = []
        warnings: list[str] = []

        for req in self.REQUIRED_SHEETS:
            if req not in data:
                errors.append(f"Brak wymaganego arkusza: '{req}'")

        if errors:
            return XlsxImportResult(success=False, graph=None, errors=errors)

        buses = data.get("Szyny", [])
        lines = data.get("Linie", [])
        trafos = data.get("Trafo", [])
        sources = data.get("Źródła", [])
        loads = data.get("Odbiory", [])

        # Validate cross-references
        bus_ids = {b["id"] for b in buses}

        for i, line in enumerate(lines):
            if line.get("szyna_pocz") not in bus_ids:
                errors.append(
                    f"Linie wiersz {i + 2}: szyna_pocz nie istnieje w Szyny"
                )
            if line.get("szyna_kon") not in bus_ids:
                errors.append(
                    f"Linie wiersz {i + 2}: szyna_kon nie istnieje w Szyny"
                )

        # Check numeric values
        for i, bus in enumerate(buses):
            if bus.get("napięcie_kV", 0) <= 0:
                errors.append(f"Szyny wiersz {i + 2}: napięcie_kV musi być > 0")

        for i, line in enumerate(lines):
            if line.get("długość_km", 0) <= 0:
                errors.append(f"Linie wiersz {i + 2}: długość_km musi być > 0")

        # Check duplicate IDs
        all_bus_ids = [b["id"] for b in buses]
        if len(all_bus_ids) != len(set(all_bus_ids)):
            errors.append("Duplikaty ID w arkuszu Szyny")

        if errors:
            return XlsxImportResult(success=False, graph=None, errors=errors)

        graph = self._build_graph(buses, lines, trafos, sources, loads, warnings)

        return XlsxImportResult(
            success=True,
            graph=graph,
            bus_count=len(buses),
            branch_count=len(lines) + len(trafos),
            source_count=len(sources),
            load_count=len(loads),
            trafo_count=len(trafos),
            warnings=warnings,
        )

    def _parse_sheet(
        self,
        sheet: Any,
        columns: dict[str, type],
        sheet_name: str,
        errors: list[str],
    ) -> list[dict[str, Any]]:
        """Parse a sheet into list of dicts with type validation."""
        rows = list(sheet.iter_rows(min_row=1, values_only=True))
        if not rows:
            errors.append(f"Arkusz '{sheet_name}' jest pusty")
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Check required columns
        for col_name in columns:
            if col_name not in headers:
                errors.append(
                    f"Arkusz '{sheet_name}': brak wymaganej kolumny '{col_name}'"
                )

        if any(col_name not in headers for col_name in columns):
            return []

        col_indices = {name: headers.index(name) for name in columns}
        result: list[dict[str, Any]] = []

        for row_num, row in enumerate(rows[1:], start=2):
            record: dict[str, Any] = {}
            row_valid = True

            for col_name, col_type in columns.items():
                idx = col_indices[col_name]
                value = row[idx] if idx < len(row) else None

                if value is None or str(value).strip() == "":
                    errors.append(
                        f"Arkusz '{sheet_name}' wiersz {row_num}: "
                        f"pusta wartość w kolumnie '{col_name}'"
                    )
                    row_valid = False
                    continue

                try:
                    record[col_name] = col_type(value)
                except (ValueError, TypeError):
                    errors.append(
                        f"Arkusz '{sheet_name}' wiersz {row_num}: "
                        f"nieprawidłowa wartość '{value}' w kolumnie '{col_name}' "
                        f"(oczekiwano {col_type.__name__})"
                    )
                    row_valid = False

            if row_valid:
                result.append(record)

        return result

    def _build_graph(
        self,
        buses: list[dict[str, Any]],
        lines: list[dict[str, Any]],
        trafos: list[dict[str, Any]],
        sources: list[dict[str, Any]],
        loads: list[dict[str, Any]],
        warnings: list[str],
    ) -> NetworkGraph:
        """Build NetworkGraph from parsed data."""
        graph = NetworkGraph()

        # Determine slack bus (first source bus, or first bus)
        source_bus_ids = {s["szyna"] for s in sources} if sources else set()

        # Build a lookup for loads by bus id for PQ node parameters
        load_by_bus: dict[str, dict[str, float]] = {}
        for load in loads:
            bus_id = load["szyna"]
            if bus_id not in load_by_bus:
                load_by_bus[bus_id] = {"P_MW": 0.0, "Q_Mvar": 0.0}
            load_by_bus[bus_id]["P_MW"] += load["P_MW"]
            load_by_bus[bus_id]["Q_Mvar"] += load["Q_Mvar"]

        # Track whether we already assigned a SLACK node
        slack_assigned = False

        for bus in buses:
            bus_id = bus["id"]
            voltage_kv = bus["napięcie_kV"]

            if bus_id in source_bus_ids and not slack_assigned:
                node_type = NodeType.SLACK
                slack_assigned = True
                node = Node(
                    id=bus_id,
                    name=bus["nazwa"],
                    voltage_level=voltage_kv,
                    node_type=node_type,
                    voltage_magnitude=1.0,
                    voltage_angle=0.0,
                )
            else:
                node_type = NodeType.PQ
                bus_load = load_by_bus.get(bus_id, {"P_MW": 0.0, "Q_Mvar": 0.0})
                node = Node(
                    id=bus_id,
                    name=bus["nazwa"],
                    voltage_level=voltage_kv,
                    node_type=node_type,
                    active_power=-bus_load["P_MW"],
                    reactive_power=-bus_load["Q_Mvar"],
                )
            graph.add_node(node)

        # Add lines as LineBranch instances
        for line in lines:
            length_km = line["długość_km"]
            r_ohm_per_km = line["R_ohm_km"]
            x_ohm_per_km = line["X_ohm_km"]
            b_us_per_km = line.get("B_uS_km", 0.0)

            branch = LineBranch(
                id=line["id"],
                name=line.get("typ", line["id"]),
                branch_type=BranchType.LINE,
                from_node_id=line["szyna_pocz"],
                to_node_id=line["szyna_kon"],
                r_ohm_per_km=r_ohm_per_km,
                x_ohm_per_km=x_ohm_per_km,
                b_us_per_km=b_us_per_km,
                length_km=length_km,
                rated_current_a=1.0,  # placeholder — XLSX does not provide Irated
            )
            graph.add_branch(branch)

        # Add transformers as TransformerBranch instances
        for trafo in trafos:
            hv_bus = next((b for b in buses if b["id"] == trafo["szyna_HV"]), None)
            lv_bus = next((b for b in buses if b["id"] == trafo["szyna_LV"]), None)

            if hv_bus and lv_bus:
                sn_mva = trafo["Sn_MVA"]
                uk_pct = trafo["uk_pct"]
                pk_kw = trafo.get("Pk_kW", 0.0)
                u_hv = hv_bus["napięcie_kV"]
                u_lv = lv_bus["napięcie_kV"]
                vector_group = trafo.get("grupa", "Dyn11")

                branch = TransformerBranch(
                    id=trafo["id"],
                    name=vector_group,
                    branch_type=BranchType.TRANSFORMER,
                    from_node_id=trafo["szyna_HV"],
                    to_node_id=trafo["szyna_LV"],
                    rated_power_mva=sn_mva,
                    voltage_hv_kv=u_hv,
                    voltage_lv_kv=u_lv,
                    uk_percent=uk_pct,
                    pk_kw=pk_kw,
                    vector_group=vector_group,
                )
                graph.add_branch(branch)

        # Store source impedance metadata on SLACK node(s)
        for src in sources:
            bus = next((b for b in buses if b["id"] == src["szyna"]), None)
            if bus:
                sk_mva = src["Sk_MVA"]
                rx_ratio = src["RX_ratio"]
                u_kv = bus["napięcie_kV"]

                if sk_mva > 0:
                    z_total = (u_kv**2) / sk_mva
                    r_source = z_total * rx_ratio / math.sqrt(1 + rx_ratio**2)
                    x_source = z_total / math.sqrt(1 + rx_ratio**2)

                    # Attach source impedance as dynamic attribute on the node
                    node = graph.nodes[src["szyna"]]
                    node.source_impedance = complex(r_source, x_source)  # type: ignore[attr-defined]

        # Note: loads are already reflected in PQ node parameters above
        if loads:
            warnings.append(
                f"Zaimportowano {len(loads)} odbiorów "
                f"(dane dostępne w parametrach węzłów PQ)"
            )

        return graph
